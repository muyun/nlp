# -*- coding: utf-8 -*-
#  
"""
   simptext.dt_sent
   ~~~~~~~~~~
   about data processing of the sentences 

@author wenlong

@TODO: update the software construction

"""

import sys, re, codecs

from bs4 import BeautifulSoup

from collections import OrderedDict
import openpyxl, json, csv

from nltk.tokenize import StanfordTokenizer
#from nltk.tokenize import wordpunct_tokenize
from nltk.tag import StanfordPOSTagger
eng_tagger = StanfordPOSTagger('english-bidirectional-distsim.tagger')
# use the wrapper or use the standard lib?
from nltk.parse.stanford import StanfordDependencyParser
eng_parser = StanfordDependencyParser(model_path=u'edu/stanford/nlp/models/lexparser/englishPCFG.ser.gz')

# for Roget's Thesaurus (1911)
#from utils.algs import punct, coordi, subordi, adverb, parti, adjec, appos, passive, paratax

#the algs about the syntax
from algs import base, punct, coordi, subordi, adverb, parti, adjec, appos, passive, paratax

# the models

#reload(sys)
#sys.setdefaultencoding('utf-8')

def read_xlsx_file(filename, sheetnums, columnnum):
    """read the xlsx file and stored first sheetnums into words list"""
    # using the openpyxl lib here
    wb = openpyxl.load_workbook(filename)

    #import pdb; pdb.set_trace()
    sheet_names = wb.get_sheet_names()[0:sheetnums]
    #sheet1 = wb.get_sheet_by_name('level 1')
    # sheet2 = wb.get_sheet_by_name('level 2')
    #sheet = wb.get_sheet_by_name(sheets_names)

    # store the simplied words in the words list
    words = []
    for sheet_name in sheet_names:
        worksheet = wb.get_sheet_by_name(sheet_name)
        for x in range(1, worksheet.max_row+1):
            words.append(str(worksheet.cell(row=x, column=columnnum).value).lower())

    # now removing it
    # TODO- the replace function

    #import pdb; pdb.set_trace()
    return tuple(words)


def read_xml_file(filename, word):
    """return the lemmas for the word in filename"""
    lemmas = []
    soup = BeautifulSoup(open(filename))

    # import pdb; pdb.set_trace()
    tokens = soup.find_all("token")
    for tk in tokens:
        # print tk
        if tk['lemma'] == word:
            lemmas.append(tk['lemma'])
            for st in tk.find_all("subst"):
                # print st['lemma']
                lemmas.append(st['lemma'])
    return tuple(lemmas)


def get_edblist(fin): #load EDB_List
    flist = open(fin,'r')
    edb_list = []
    ltines = flist.readlines()
    for l in ltines:
        if l:
            edb_list.append(l.strip())
        else:
            break
    return edb_list

 
def get_stat_info(filename, store_filename):
    """ read the filename and store the words with lemmas in store_filename"""

    num_words_syns = 0

    docs = OrderedDict() # store the info - docs[sentence] = [id,...]
    
    lemmas = {}  # store the word info - lemmas[id] = {word: synsets, ...}
    
    soup = BeautifulSoup(open(filename), "lxml")

    # number of sentences, based on the 'sent' tag
    sentences = soup.find_all("sent")
    num_sentences = len(sentences)

    # number of tokens, based on the 'token' tag
    num_tokens = len(soup.find_all("token"))

    #import pdb; pdb.set_trace()
    for sentence in sentences:
        # the key in docs is the target sentence
        target = str(sentence.targetsentence)
        docs[target] = []
        for tk in sentence.tokens.find_all("token"):
            #print(tk)
            if tk['id'].isdigit():
                num_words_syns += 1

                docs[target].append(tk['id'])

                #test_token[tk['id']]=tk['wordform']
                #_test_token.append(tk['wordform'])
                # the same word are stored in the same slot in the lemmas dict; A better data structure should be used
                
                lemmas[tk['id']] = []
                lemmas[tk['id']].append(tk['wordform'])
                lemmas[tk['id']].append(tk['lemma'])
                for st in tk.find_all("subst"):
                    lemmas[tk['id']].append(st['lemma'])

    #print "#sentence: ", num_sentence
    #print "#words: ", num_words
    #print "#words_syns: ", num_words_syns
    #print "#_test_token: ", len(_test_token)
    #print "#lemmas: ", len(lemmas)

    """
    for lst in _test_token:
        if lst in lemmas:
            pass
        else:
            print "lst: ", lst
    """

    """
    import pdb; pdb.set_trace()
    k = lemmas.keys()      
    k_feas = list(set(k) - set(_test_token))
    print(k_feas)                   
    """
    # write the file

    #write the fiel
    json.dump(lemmas, open(store_filename, 'w'))
    
    return num_sentences, num_tokens, num_words_syns, docs          


def print_intermedia(datafile, docs, wordlist):
    """print the intermedia data for the check
     @datafile is the lemmas dict filename
     @docs is the targetsentence
     @wordlist is the edb wordlist

    """
    data = json.load(open(datafile))

    # import pdb; pdb.set_trace()
    for sentence in docs.keys():
        #print(sentence)
        #print(docs[sentence])
        
        output = OrderedDict()
        for id in docs[sentence]:
            w = data[id][1] # the lemma of the wordform
            # remove the original word
            coincolist = data[id][1:]
            if w in coincolist:
                coincolist.remove(w)
            
            wordlist = wordcal.get_wordnet_list(w)
            if w in wordlist:
                wordlist.remove(w)

            feas = set(coincolist).intersection(wordlist)
            if len(feas) >= 1:
                k = '*'+data[id][0]
                output[k] = [coincolist, wordlist]
            else:
                output[data[id][0]] = [coincolist, wordlist]

        # write the sentence

        #import pdb; pdb.set_trace()

        with open('coinco_l1.json', 'a') as outfile:
            outfile.write('\n'+sentence+'\n')
            json.dump(output, outfile, indent=2)

    #json.dump(output, open('intermedia.json', 'w'))

def print_syn_sent(filename, sent_file):
    num_sentences = 0
    num_splitted_sentences = 0
    #data = json.load(open(datafile))
    docs = OrderedDict() # store the info - docs[sentence] = [id,...]
    
    soup = BeautifulSoup(open(filename), "lxml")

    # number of sentences, based on the 'sent' tag
    sentences = soup.find_all("targetsentence")
    #num_sentences = len(sentences)

    p = re.compile(r'<.*?>')
    #import pdb; pdb.set_trace()
    output = OrderedDict()
    for sentence in sentences:
        num_sentences = num_sentences + 1
        #print(sentence)
        sent = str(p.sub('', str(sentence)))
        se = re.sub(r'^”|”$', '', sent)
        #sent = str(BeautifulSoup(sentence).text)
        print(se)
        # write the sentence
        #res = ""
        res = alg.simp_coordi_sent(se)
        #res = alg.simp_subordi_sent(se)
        #res = alg.simp_advcl_sent(se)
        #res = alg.simp_parti_sent(se)
        if res: # the
            num_splitted_sentences = num_splitted_sentences + 1
            
        
        output[sentence] = res
        #import pdb; pdb.set_trace()
        
        with open(sent_file, 'a') as outfile:
           outfile.write(str(sentence)+'\n')
           outfile.write("OUTPUT: " + res + '\n')
           outfile.write('-----------------------\n')
           #json.dump(output, outfile, indent=2)

    return num_sentences, num_splitted_sentences       


def print_semeval_sent(filename, sent_file):
    num_sentences = 0
    num_splitted_sentences = 0
    #data = json.load(open(datafile))
    docs = OrderedDict() # store the info - docs[sentence] = [id,...]
    
    soup = BeautifulSoup(open(filename), "lxml")

    # number of sentences, based on the 'sent' tag
    sentences = soup.find_all("instance")
    #num_sentences = len(sentences)

    p = re.compile(r'<.*?>')
    #import pdb; pdb.set_trace()
    output = OrderedDict()
    for sentence in sentences:

        #import pdb; pdb.set_trace()
        num_sentences = num_sentences + 1
        #print(sentence)
        #sent = str(p.sub('', str(sentence)))
        #se = re.sub(r'^”|”$', '', sent)
        se = sentence.context.get_text()
        #sent = str(BeautifulSoup(sentence).text)
        #print(se)
        # write the sentence
        #res = ""
        res = alg.simp_coordi_sent(se)
        #res = alg.simp_subordi_sent(se)
        #res = alg.simp_advcl_sent(se)
        #res = alg.simp_parti_sent(se)
        if res: # the
            num_splitted_sentences = num_splitted_sentences + 1
            
        output[sentence] = res
        #import pdb; pdb.set_trace()
        
        with open(sent_file, 'a') as outfile:
           outfile.write(str(sentence)+'\n')
           outfile.write("OUTPUT: " + res + '\n')
           outfile.write('-----------------------\n')
           #json.dump(output, outfile, indent=2)

    return num_sentences, num_splitted_sentences

def print_mturk_sent(filename, sent_file):
    num_sentences = 0
    num_splitted_sentences = 0
    #data = json.load(open(datafile))
    docs = OrderedDict() # store the info - docs[sentence] = [id,...]
    
    #soup = BeautifulSoup(open(filename), "lxml")

    # number of sentences, based on the 'sent' tag
    #sentences = soup.find_all("instance")
    #num_sentences = len(sentences)

    #p = re.compile(r'<.*?>')
    #import pdb; pdb.set_trace()
    output = OrderedDict()
    f = open(filename, 'rU')
    num = 0
    res = ""
    for line in f:
        line = line.strip('\n')
        if line:
            #import pdb; pdb.set_trace()
            num_sentences = num_sentences + 1
            #print(sentence)
            #sent = str(p.sub('', str(sentence)))
            #se = re.sub(r'^”|”$', '', sent)
            #se = sentence.context.get_text()
            #sent = str(BeautifulSoup(sentence).text)
            obj = line.split("\t")
            se = re.sub(r'^"|"$', '', obj[0])
            print(se)
            # write the sentence
            #res = ""
            #res = punct.simp_syn_sent_(se)
            #res = coordi.simp_syn_sent_(se)
            #res = subordi.simp_syn_sent_(str(se))
            #res = adverb.simp_syn_sent_(str(se))
            #res = parti.simp_syn_sent_(str(se))
            #res = adjec.simp_syn_sent_(str(se))
            #res = appos.simp_syn_sent_(str(se))
            #res = passive.simp_syn_sent_(str(se))
            #res = paratax.simp_syn_sent_(str(se))
            #res = alg.simp_passive_sent(str(re))
            _res = simp_syn_sent(str(se))
            if len(_res)>0:            
                (s1, s1_child, s2, s2_child, res) = _get_split_ret(_res)
                print "res: ", res
            else:
                res = _res

            if _res: # the
                num_splitted_sentences = num_splitted_sentences + 1

            #import pdb; pdb.set_trace()
            output[se] = res
            #import pdb; pdb.set_trace()
        
            with open(sent_file, 'a') as outfile:
                outfile.write(str(se)+'\n')
                outfile.write("OUTPUT: " + res + '\n')
                #outfile.write('-----------------------\n')
                #json.dump(output, outfile, indent=2)

            """  
            num = num + 1
            if num == 200:
                break
            """

    return num_sentences, num_splitted_sentences


def cal_mturk_sent(filename, gt):
    f = open(filename, 'rU')
    _num_output = 0
    num_negative = 0
    num_false_positive = 0
    num_true_negative = 0
    num_false_negative = 0
    num_true_positive = 0
    _num_false_positive = 0
    _num_true_negative = 0
    _num_false_negative = 0
    _num_true_positive = 0
    num_positive = 0
    num = 0
    len_input = 0
    len_output = 0
    _input = ""
    output = OrderedDict()
    for line in f:
        line = line.strip('\n')

        #import pdb; pdb.set_trace()
        if "OUTPUT" not in line :
            match = re.search(r'--+', line)
            if match:
                pass
            else:
                _input = line
                len_input = len(re.findall("\w+", line))
            #print "Input: ", line    

        #import pdb; pdb.set_trace()
        else:
            #print "gt-out: ", gt[num]
            #print "out: ", line
            len_output = len(re.findall("\w+", line))
            #import pdb; pdb.set_trace()
            #num_output = num_output + 1
            #print(sentence)
            #sent = str(p.sub('', str(sentence)))
            #se = re.sub(r'^”|”$', '', sent)
            #se = sentence.context.get_text()
            #sent = str(BeautifulSoup(sentence).text)
            ot_flag = line.split(':')[0]
            ot = line.split(':')[1].strip() # having output or not

            #import pdb; pdb.set_trace()
            if ot_flag == "OUTPUT":  # the num
                _num_output = _num_output + 1

            if not ot: #don't have the output,consider it as false negative
                num_negative = num_negative + 1
                _sp = ['N']
                
                if gt[num] != 'x':
                    num_false_negative = num_false_negative + 1
                    #_sp = ['N']
                else:
                    num_true_negative = num_true_negative + 1
                    #_sp = ['N']

                if gt[num] != 'x':
                    _num_false_negative = _num_false_negative + 1
                    _gen = ['N']
                else:
                    _num_false_positive = _num_false_positive + 1
                    _gen =['N']
                    
            else: # having the output
                num_positive = num_positive + 1
                _sp = ['Y']

                if gt[num] == 'x': #False positive
                    num_false_positive = num_false_positive + 1
                    #_sp = ['Y']
                    
                else:
                    num_true_positive = num_true_positive + 1
                    #_sp = ['Y']
                
                if abs(len(re.findall("\w+", gt[num])) - len_output) <= 5 :
                    _num_true_positive = _num_true_positive + 1
                    _gen = ['Y']
                else:
                    _num_false_positive = _num_false_positive + 1
                    _gen =['N']

            
            output[num] = [_input,
                         gt[num],
                         ot,
                         _sp,
                        _gen]
            
            #import pdb; pdb.set_trace()
            with codecs.open('mturk_sent_l5.csv', 'a', encoding='utf-8') as outfile:
                wr = csv.writer(outfile, delimiter = ',', quoting = csv.QUOTE_ALL)
                wr.writerow(output[num])

            num = num + 1
            if num == 300:
                break    

    #import pdb; pdb.set_trace()
            #match = re.search(r'(^(#OUTPUT):(\w*))', line)
    print "#num_negative: ", num_negative
    print "#num_false_positive: ", num_false_positive
    print "#num_true_positive: ", num_true_positive
    print "#num_false_negative: ", num_false_negative
    print "#num_true_negative: ", num_true_negative
    print "#num_positive: ", num_positive
    print "#_num_output: ", _num_output
    
    return num_negative, num_false_positive, num_false_negative, num_positive       


def simp_mturk_sent(filename, sent_file):
    num_sentences = 0
    num_splitted_sentences = 0
    #data = json.load(open(datafile))
    docs = OrderedDict() # store the info - docs[sentence] = [id,...]
    
    #soup = BeautifulSoup(open(filename), "lxml")

    # number of sentences, based on the 'sent' tag
    #sentences = soup.find_all("instance")
    #num_sentences = len(sentences)

    #p = re.compile(r'<.*?>')
    #import pdb; pdb.set_trace()
    output = OrderedDict()
    _output = OrderedDict()
    f = open(filename, 'rU')
    num = 0
    for line in f:
        line = line.strip('\n')
        if line:
            #import pdb; pdb.set_trace()
            num_sentences = num_sentences + 1
            #print(sentence)
            #sent = str(p.sub('', str(sentence)))
            #se = re.sub(r'^”|”$', '', sent)
            #se = sentence.context.get_text()
            #sent = str(BeautifulSoup(sentence).text)
            obj = line.split("\t")
            se = re.sub(r'^"|"$', '', obj[0])
            #print(se)
            # write the sentence
            #res = ""
            #res = punct.simp_syn_sent_(se)
            #res = coordi.simp_syn_sent_(se)
            #res = subordi.simp_syn_sent_(str(se))
            #res = adverb.simp_syn_sent_(str(se))
            #res = parti.simp_syn_sent_(str(se))
            #res = adjec.simp_syn_sent_(str(se))
            #res = appos.simp_syn_sent_(str(se))
            #res = passive.simp_syn_sent_(str(se))
            #res = paratax.simp_syn_sent_(str(se))
            #res = alg.simp_passive_sent(str(re))
            _res = simp_syn_sent(str(se))

            if len(_syn_ret)>0:
                res = _get_split_ret(_res)
            #import pdb; pdb.set_trace()
            #print "res: ", res
            if res:  # the
                num_splitted_sentences = num_splitted_sentences + 1

            #import pdb; pdb.set_trace()
            output[se] = res
            #import pdb; pdb.set_trace()
        
            with open(sent_file, 'a') as outfile:
                outfile.write(str(se)+'\n')
                outfile.write("OUTPUT: " + res + '\n')
                outfile.write('-----------------------\n')
                #json.dump(output, outfile, indent=2)

            _output[se] = [se, res]
            with codecs.open('mturk_sent_.csv', 'a') as _outfile:
                wr = csv.writer(_outfile, delimiter = ',', quoting = csv.QUOTE_ALL)
                wr.writerow(_output[se])
                
            """
            num = num + 1
            if num == 300:
                break
            """  
            
    return num_sentences, num_splitted_sentences        


def simp_semeval_sent(filename, sent_file):
    num_sentences = 0
    num_splitted_sentences = 0
    #data = json.load(open(datafile))
    # docs = OrderedDict() # store the info - docs[sentence] = [id,...]
    
    #soup = BeautifulSoup(open(filename), "lxml")

    # number of sentences, based on the 'sent' tag
    #sentences = soup.find_all("instance")
    #num_sentences = len(sentences)

    #p = re.compile(r'<.*?>')
    #import pdb; pdb.set_trace()
    output = OrderedDict()
    _output = OrderedDict()
    soup = BeautifulSoup(open(filename), "lxml")

    """
    num_sentences = 0
    num_tokens = 0
    num_words_syns = 0
    """

    num = 0
    # semeval = OrderedDict()  # semeval -> {side.n 301 : [sent], [synsets]}
    
    lexelts = soup.find_all("lexelt")
    for lex in lexelts:
        # num_words_syns = num_words_syns + 1
        item = lex["item"]
        for sentence in lex.find_all("instance"):
            num_sentences = num_sentences + 1

            id = sentence["id"]
            key = item + " " + id
            se = sentence.context.get_text()

            # import pdb; pdb.set_trace()
            # words = sent.split()
            # num_tokens = num_tokens + len(words) - 1

            """
            semeval[key] = []
            semeval[key].append(sent)
            if key in synsets:
                # print(synsets[key])
                semeval[key].append(synsets[key])
            """    
  
            #import pdb; pdb.set_trace()
            num_sentences = num_sentences + 1
            #print(sentence)
            #sent = str(p.sub('', str(sentence)))
            #se = re.sub(r'^”|”$', '', sent)
            #se = sentence.context.get_text()
            #sent = str(BeautifulSoup(sentence).text)
            #obj = line.split("\t")
            #se = re.sub(r'^"|"$', '', obj[0])
            #print(se)
            # write the sentence
            #res = ""
            #res = punct.simp_syn_sent_(se)
            #res = coordi.simp_syn_sent_(se)
            #res = subordi.simp_syn_sent_(str(se))
            #res = adverb.simp_syn_sent_(str(se))
            #res = parti.simp_syn_sent_(str(se))
            #res = adjec.simp_syn_sent_(str(se))
            #res = appos.simp_syn_sent_(str(se))
            #res = passive.simp_syn_sent_(str(se))
            #res = paratax.simp_syn_sent_(str(se))
            #res = alg.simp_passive_sent(str(re))
            res = simp_syn_sent(str(se))
            #import pdb; pdb.set_trace()
            #print "res: ", res
            if res:  # the
                num_splitted_sentences = num_splitted_sentences + 1

            #import pdb; pdb.set_trace()
            output[se] = res
            #import pdb; pdb.set_trace()
        
            with open(sent_file, 'a') as outfile:
                outfile.write(str(se)+'\n')
                outfile.write("OUTPUT: " + res + '\n')
                outfile.write('-----------------------\n')
                #json.dump(output, outfile, indent=2)

            _output[se] = [se, res]
            with codecs.open('semeval_sent.csv', 'a') as _outfile:
                wr = csv.writer(_outfile, delimiter = ',', quoting = csv.QUOTE_ALL)
                wr.writerow(_output[se])
            """    
            num = num + 1
            if num == 20:
                break
            """
              
    return num_sentences, num_splitted_sentences    


def simp_semeval_sent(filename, sent_file):
    num_sentences = 0
    num_splitted_sentences = 0
    #data = json.load(open(datafile))
    #docs = OrderedDict() # store the info - docs[sentence] = [id,...]
    
    #soup = BeautifulSoup(open(filename), "lxml")

    # number of sentences, based on the 'sent' tag
    #sentences = soup.find_all("instance")
    #num_sentences = len(sentences)

    #p = re.compile(r'<.*?>')
    #import pdb; pdb.set_trace()
    output = OrderedDict()
    _output = OrderedDict()
    soup = BeautifulSoup(open(filename), "lxml")

    # num_sentences = 0
    # num_tokens = 0

    #semeval = OrderedDict()  # semeval -> {side.n 301 : [sent], [synsets]}
    num = 0
    lexelts = soup.find_all("lexelt")
    for lex in lexelts:
        # num_words_syns = num_words_syns + 1
        item = lex["item"]
        for sentence in lex.find_all("instance"):
            num_sentences = num_sentences + 1

            # id = sentence["id"]
            # key = item + " " + id
            se = sentence.context.get_text()

            # import pdb; pdb.set_trace()
            #words = sent.split()
            #num_tokens = num_tokens + len(words) - 1

            """
            semeval[key] = []
            semeval[key].append(se)
            if key in synsets:
                # print(synsets[key])
                semeval[key].append(synsets[key])
            """

            #import pdb; pdb.set_trace()
            num_sentences = num_sentences + 1
            #print(sentence)
            #sent = str(p.sub('', str(sentence)))
            #se = re.sub(r'^”|”$', '', sent)
            #se = sentence.context.get_text()
            #sent = str(BeautifulSoup(sentence).text)
            #obj = line.split("\t")
            #se = re.sub(r'^"|"$', '', obj[0])
            #print(se)
            # write the sentence
            #res = ""
            #res = punct.simp_syn_sent_(se)
            #res = coordi.simp_syn_sent_(se)
            #res = subordi.simp_syn_sent_(str(se))
            #res = adverb.simp_syn_sent_(str(se))
            #res = parti.simp_syn_sent_(str(se))
            #res = adjec.simp_syn_sent_(str(se))
            #res = appos.simp_syn_sent_(str(se))
            #res = passive.simp_syn_sent_(str(se))
            #res = paratax.simp_syn_sent_(str(se))
            #res = alg.simp_passive_sent(str(re))
            res = simp_syn_sent(str(se))
            #import pdb; pdb.set_trace()
            #print "res: ", res
            if res:  # the
                num_splitted_sentences = num_splitted_sentences + 1

            #import pdb; pdb.set_trace()
            output[se] = res
            #import pdb; pdb.set_trace()
        
            with open(sent_file, 'a') as outfile:
                outfile.write(str(se)+'\n')
                outfile.write("OUTPUT: " + res + '\n')
                outfile.write('-----------------------\n')
                #json.dump(output, outfile, indent=2)

            _output[se] = [se, res]
            with codecs.open('semeval_sent.csv', 'a') as _outfile:
                wr = csv.writer(_outfile, delimiter = ',', quoting = csv.QUOTE_ALL)
                wr.writerow(_output[se])
                
            num = num + 1
            if num == 20:
                break
               
            
    return num_sentences, num_splitted_sentences    

def simp_coinco_sent(filename, sent_file):
    num_sentences = 0
    num_splitted_sentences = 0
    #data = json.load(open(datafile))
    #docs = OrderedDict() # store the info - docs[sentence] = [id,...]
    
    docs = OrderedDict()  # store the info - docs[sentence] = [id,...]
    output = OrderedDict()
    _output = OrderedDict()
    soup = BeautifulSoup(open(filename), "lxml")

    # number of sentences, based on the 'sent' tag
    sentences = soup.find_all("targetsentence")
    # num_sentences = len(sentences)
    num = 0
    # number of tokens, based on the 'token' tag
    # num_tokens = len(soup.find_all("token"))

    p = re.compile(r'<.*?>')
    # import pdb; pdb.set_trace()
    #output = OrderedDict()
    for sentence in sentences:
        # print(sentence)
        se = str(p.sub('', str(sentence)))
        num_sentences = num_sentences + 1

            #print(sentence)
            #sent = str(p.sub('', str(sentence)))
            #se = re.sub(r'^”|”$', '', sent)
            #se = sentence.context.get_text()
            #sent = str(BeautifulSoup(sentence).text)
            #obj = line.split("\t")
            #se = re.sub(r'^"|"$', '', obj[0])
            #print(se)
            # write the sentence
            #res = ""
            #res = punct.simp_syn_sent_(se)
            #res = coordi.simp_syn_sent_(se)
            #res = subordi.simp_syn_sent_(str(se))
            #res = adverb.simp_syn_sent_(str(se))
            #res = parti.simp_syn_sent_(str(se))
            #res = adjec.simp_syn_sent_(str(se))
            #res = appos.simp_syn_sent_(str(se))
            #res = passive.simp_syn_sent_(str(se))
            #res = paratax.simp_syn_sent_(str(se))
            #res = alg.simp_passive_sent(str(re))
        res = simp_syn_sent(str(se))
            #import pdb; pdb.set_trace()
            #print "res: ", res
        if res:  # the
            num_splitted_sentences = num_splitted_sentences + 1

            #import pdb; pdb.set_trace()
        output[se] = res
            #import pdb; pdb.set_trace()
        
        with open(sent_file, 'a') as outfile:
            outfile.write(str(se)+'\n')
            outfile.write("OUTPUT: " + res + '\n')
            outfile.write('-----------------------\n')
                #json.dump(output, outfile, indent=2)

        _output[se] = [se, res]
        with codecs.open('coinco_sent.csv', 'a') as _outfile:
            wr = csv.writer(_outfile, delimiter = ',', quoting = csv.QUOTE_ALL)
            wr.writerow(_output[se])
        
        """        
        num = num + 1
        if num == 20:
            break
        """                
    return num_sentences, num_splitted_sentences    


def simp_syn_sent(sent):
    strs = ""
    # the original tokens in the sent
    #import pdb; pdb.set_trace()
    #print "syn sent: ", sent
    #import pdb; pdb.set_trace()
    tokens = StanfordTokenizer().tokenize(sent)
    #tokens = wordpunct_tokenize(strs)
    tokens.insert(0, '')

    taggers = eng_tagger.tag(sent.split())

    result = list(eng_parser.raw_parse(sent))[0]
    root = result.root['word']

    #w = result.tree()
    #print "parse_tree:", w
    
    #TODO: use the tree structure, Check again
    node_list = [] # dict (4 -> 4, u'said', u'VBD', u'root', [[18], [22], [16], [3]])
    for node in result.nodes.items():
        node_list.append(base.get_triples(node))
        #node_list[base.get_triples[0]] = base.get_triples(node)

    #import pdb; pdb.set_trace()
    if len(sent) > 0:
        strs = paratax.simp_paratax_sent(tokens, node_list)
        if len(strs) > 0:
            return strs
        else:
            strs = punct.simp_punct_sent(tokens, taggers, node_list)
            if len(strs) > 0:
                return strs
            else:               
                strs = coordi.simp_coordi_sent(tokens, node_list)
                if len(strs) > 0:
                    return strs
                else:        
                    strs = subordi.simp_subordi_sent(tokens, node_list)
                    if len(strs) > 0:
                        return strs
                    else:
                        strs = adverb.simp_adverb_sent(tokens, node_list)
                        if len(strs) > 0:
                            return strs
                        else:
                            strs = passive.simp_passive_sent(tokens, node_list)
                            if len(strs) > 0:
                                return strs
                            else:
                                strs = parti.simp_parti_sent(tokens, node_list)
                                if len(strs) > 0:
                                    return strs
                                else:
                                    strs = appos.simp_appos_sent(tokens, node_list)
                                    if len(strs) > 0:
                                        return strs
                                    else:
                                        strs = adjec.simp_adjec_sent(tokens, node_list)
                                        if len(strs) > 0:
                                            return strs

    return strs 


def get_split_ret(sents):
    #
    ret = ""
    for sent in sents.split('.'):
        if len(sent) != 0:
            print "sent: ", sent
            _ret = _get_split_ret(str(sent))
            print "_ret: ", _ret
            ret = ret + str(_ret)
            print "ret: ", ret

    return ret

def _get_split_ret(_str):
    print "S1+S2: ", _str
    syn_ret = ""

    _strs = _str.split('.')

    s1 = _strs[0] + ' . '
    print "S1: ", s1
    s1_child = ""
    s2 = ""
    s2_child = ""
    #syn_ret = ""
    if len(_strs) == 1:
        return (s1, s1_child, s2, s2_child, _str)
        
    s1_child = simp_syn_sent(s1)
    print "S11+S12: ", s1_child
    
    #print "str2: ", _strs[1] 
    s2 = _strs[1] + ' .'
    print "S2: ", s2
    s2_child = simp_syn_sent(s2)
    print "S21+S22: ", s2_child

    if len(s1_child)>0: # syn_ret1
        if len(s2_child)>0: # syn_ret2
            syn_ret = s1_child + s2_child
            print "1+2: ", syn_ret
        else:
            syn_ret = s1_child + s2 
            print "1+in+2: ", syn_ret
    else:
        if len(s2_child)>0:
            syn_ret = s1 + s2_child
            print "in+1+2: ", syn_ret
        else:
            syn_ret = _str 
            print "in+1+in+2: ", syn_ret   

    print "Syntactic result: ", syn_ret

    return (s1, s1_child, s2, s2_child, syn_ret)

# Main test
def main():
    dir="/Users/zhaowenlong/workspace/proj/dev.nlp/simptext/simptext/"

   
    #filename = dir + "dataset/coinco/coinco.xml"
    #store_filename = dir + "dataset/coinco/coinco_lemmas.txt"

    """
    info = get_stat_info(filename, store_filename)
    print "#sentences: ", info[0]
    print "#words: ", info[1]
    print "#words marked with synonyms: ", info[2]
    #print "words with synonyms: ", info[3]

    #
    xlsx_filename = dir + "dataset/wordlist.xlsx"

    wordlist = read_xlsx_file(xlsx_filename, 1)
 
    info_ = get_coinco_wordlist(store_filename, wordlist)
    print "#words in the EDB list for level 1: ", info_[0]
    print "#words not in the EDB list for level 1:: ", info_[1]
    print "The ceiling for level 1: ", info_[2]
    
    
    """
    """

    # print the intermeida data
    inter = print_intermedia(store_filename, info[3], wordlist)

    """
    # print the inter data in the syntactic simplification
    #filename = dir + "utils/semeval/test/lexsub_test.xml"
    filename = dir + "utils/mturk/lex.mturk.txt"
    sent_file = dir + "tests/sent_mturk_l5_.md"
    gt_file = dir + "dataset/simplify_testset_0817.xlsx"

    # generate the output
    #_info = print_mturk_sent(filename, sent_file)
    #print "Type: Paratax Clauses:"
    #print "#sentence in mturk: ", _info[0]
    #print "#sentence with Syntactic simplification: ", _info[1]
    

    # recall and precision
    #filename = dir + "utils/testset_groundtruth.md"
    #filename = dir + "utils/coordi_mturk_l1_.json"
    #filename = dir + "utils/testset_gt_adverb.md"
    #filename = dir + "utils/testset_gt_appos.md"

    #filename = dir + "utils/testset/sent_mturk_l4_.md"
    gt = read_xlsx_file(gt_file, 1)
    _info = cal_mturk_sent(sent_file, gt)
    

    """
    lemmas = []
    wd ='mission'
    for pos in roget.parts_of_speech:

        import pdb; pdb.set_trace()
        words = roget.all_entries(wd, pos)
        for w in words:
            lemmas = lemmas + list(w)

    print lemmas
    """

    """
    filename = dir + "utils/semeval/test/lexsub_test.xml"
    sent_file = dir + "utils/testset/sent_semeval_l4_.md"
    info = simp_semeval_sent(filename, sent_file)
    print "#sentence in semeval: ", info[0]
    print "#sentence with Syntactic simplification: ", info[1]
    """
    """
    filename = dir + "utils/coinco/coinco.xml"
    sent_file = dir + "utils/testset/sent_coinco_l4_.md"
    info = simp_coinco_sent(filename, sent_file)
    print "#sentence in coinco: ", info[0]
    print "#sentence with Syntactic simplification: ", info[1]
    """


if __name__ == '__main__':
     main()
