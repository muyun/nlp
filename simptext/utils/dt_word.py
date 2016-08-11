#!/usr/bin/env python
# -*- coding: utf-8 -*-
#  
"""
   utils.dt
   ~~~~~~~~~~
   data processing

@author wenlong

@TODO: update the software construction

"""
import sys
import re
import codecs

from bs4 import BeautifulSoup

from collections import OrderedDict
import openpyxl
import json, csv

#from nltk.tag import pos_tag

# for Roget's Thesaurus (1911)
import roget, cal

import wordcal

reload(sys)
sys.setdefaultencoding('utf-8')


def read_xlsx_file(filename, sheetnums):
    """read the xlsx file and stored first sheetnums into words list"""
    # using the openpyxl lib here
    wb = openpyxl.load_workbook(filename)

    # import pdb; pdb.set_trace()
    sheet_names = wb.get_sheet_names()[0:sheetnums]
    # sheet1 = wb.get_sheet_by_name('level 1')
    # sheet2 = wb.get_sheet_by_name('level 2')
    # sheet = wb.get_sheet_by_name(sheets_names)

    # store the simplied words in the words list
    words = []
    for sheet_name in sheet_names:
        worksheet = wb.get_sheet_by_name(sheet_name)
        for x in range(1, worksheet.max_row+1):
            words.append(str(worksheet.cell(row=x, column=1).value).lower())

    # now removing it
    # TODO- the replace function

    # import pdb; pdb.set_trace()
    return tuple(words)


def read_xml_file(filename, word):
    """return the lemmas for the word in filename"""
    lemmas = []
    soup = BeautifulSoup(open(filename))

    # import pdb; pdb.set_trace()
    tokens = soup.find_all("token")
    for tk in tokens:
        # print tk
        if tk['lemma'] == word:
            lemmas.append(tk['lemma'])
            for st in tk.find_all("subst"):
                # print st['lemma']
                lemmas.append(st['lemma'])
    return tuple(lemmas)


def remove_tags(sentence):
    p = re.compile(r'<.*?>')
    sent = str(p.sub('', str(sentence)))
    se = re.sub(r'^”|”$', '', sent)

    return se

def get_stat_info(filename, store_filename):
    """ read the filename and store the words with lemmas in store_filename"""

    num_words_syns = 0

    docs = OrderedDict() # store the info - docs[sentence] = [id,...]
    lemmas = {}  # store the word info - lemmas[id] = {word: synsets, ...}
    soup = BeautifulSoup(open(filename), "lxml")

    # number of sentences, based on the 'sent' tag
    sentences = soup.find_all("sent")
    num_sentences = len(sentences)

    # number of tokens, based on the 'token' tag
    num_tokens = len(soup.find_all("token"))

    #import pdb; pdb.set_trace()
    for sentence in sentences:
        # the key in docs is the target sentence
        target = remove_tags(str(sentence.targetsentence))
        docs[target] = []
        for tk in sentence.tokens.find_all("token"):
            # print(tk)
            if tk['id'].isdigit():
                num_words_syns += 1

                docs[target].append(tk['id'])

                # test_token[tk['id']]=tk['wordform']
                # _test_token.append(tk['wordform'])
                #   the same word are stored in the same slot in the lemmas dict
                # A better data structure should be used
                lemmas[tk['id']] = []
                lemmas[tk['id']].append(tk['wordform'])
                lemmas[tk['id']].append(tk['postt'])
                lemmas[tk['id']].append(tk['lemma'])
                for st in tk.find_all("subst"):
                    _lemma = st['lemma']
                    _pos = st['pos']
                    lemmas[tk['id']].append(_lemma+"@"+_pos)

    # print "#sentence: ", num_sentence
    # print "#words: ", num_words
    # print "#words_syns: ", num_words_syns
    # print "#_test_token: ", len(_test_token)
    # print "#lemmas: ", len(lemmas)

    """
    for lst in _test_token:
        if lst in lemmas:
            pass
        else:
            print "lst: ", lst
    """

    """
    import pdb; pdb.set_trace()
    k = lemmas.keys()      
    k_feas = list(set(k) - set(_test_token))
    print(k_feas)                   
    """
    # write the file

    # write the fiel
    json.dump(lemmas, open(store_filename, 'w'))
    return num_sentences, num_tokens, num_words_syns, docs          


def print_intermedia(datafile, docs, wordlist):
    """print the intermedia data for the check
     @datafile is the lemmas dict filename
     @docs is the targetsentence
     @wordlist is the edb wordlist

    """
    data = json.load(open(datafile))

    #import pdb; pdb.set_trace()
    for sentence in docs.keys():
        # print(sentence)
        # print(docs[sentence])
        
        output = OrderedDict()
        for id in docs[sentence]:
            w = data[id][2]  # the lemma of the wordform
            # remove the original word
            pos = data[id][1] # pos of the wordform
            pos_coincolist = data[id][3:]

            coincolist = []
            for _pos in pos_coincolist:
                coincolist.append(_pos.split('@')[0])
                
            if w in coincolist:
                coincolist.remove(w)
            
            wordlist = wordcal.get_wordnet_list(w)
            #wordlist = wordcal.get_hyponyms(w)
            if w in wordlist:
                wordlist.remove(w)

            pos_wordlist = []    
            for w in wordlist:
                #import pdb; pdb.set_trace()
                _w = wordcal.get_word_pos(str(w))

                pos_wordlist.append(_w)

            feas = set(coincolist).intersection(wordlist)
            if len(feas) >= 1 :
                k = '*'+data[id][0]
                wf = ['Y']
            else:
                k = data[id][0]
                wf = ['N']

            output[k] = [sentence,
                         k+'@'+pos,
                         pos_coincolist,
                         pos_wordlist,
                         wf]    

        # write the sentence

            #import pdb; pdb.set_trace()
            with codecs.open('coinco_l1.csv', 'a') as outfile:
                wr = csv.writer(outfile, delimiter = ',', quoting = csv.QUOTE_ALL)
                #for row in output:
                wr.writerow(output[k])    

    # json.dump(output, open('intermedia.json', 'w'))


def print_coinco_sent(filename):
    # data = json.load(open(datafile))
    docs = OrderedDict()  # store the info - docs[sentence] = [id,...]
    
    soup = BeautifulSoup(open(filename), "lxml")

    # number of sentences, based on the 'sent' tag
    sentences = soup.find_all("targetsentence")
    # num_sentences = len(sentences)

    # number of tokens, based on the 'token' tag
    # num_tokens = len(soup.find_all("token"))

    p = re.compile(r'<.*?>')
    # import pdb; pdb.set_trace()
    output = OrderedDict()
    for sentence in sentences:
        # print(sentence)
        sent = str(p.sub('', str(sentence)))
        # sent = str(BeautifulSoup(sentence).text)
        print(sent)
        # write the sentence
        # res = ""
        res = algs.simp_conj_sent(sent)
        output[sentence] = res
        # import pdb; pdb.set_trace()
        
        with open('coinco_conj_sent_l1_.json', 'a') as outfile:
            outfile.write(str(sentence)+'\n')
            outfile.write("OUTPUT: " + res + '\n')
            outfile.write('-----------------------\n')
            # json.dump(output, outfile, indent=2)
        

def get_coinco_wordlist(datafile, wordlist):
    #
    num_simp_words = 0  # number of words from datafile and can be found in wordlist
    num_not_simp_words = 0
    _num_not_simp_words = 0

    num_feasible_words = 0
    all_num_feasible_words = 0
    # the words with synonyms that are not in EDB list
    not_simp_wordlist = []
    
    # load the dict from coinco dataset
    data = json.load(open(datafile))

    num_words = 0
    #import pdb; pdb.set_trace()
    for id in data.keys():

        #import pdb; pdb.set_trace()
        # print data[k] # the word
        num_words = num_words + 1
        # TOUPDATE
        w = data[id][0].lower()
        # the hyponyms from wordnet of w 
        # k_wordnet_list = wordcal.get_hyponyms(w)
        k_wordnet_list = wordcal.get_wordnet_list(w)
        for w in k_wordnet_list:
            k_wordnet_list.remove(w)

        pos_coincolist = data[id][3:]
        coincolist = []
        for pos in pos_coincolist:
            coincolist.append(pos.split('@')[0])
                
        if w in coincolist:
            coincolist.remove(w)
            
        if w in wordlist: # the lemma
            num_simp_words += 1    
            
        else:
            if len(k_wordnet_list) >= 1:
                num_not_simp_words += 1
            
            not_simp_wordlist.append(data[id][0])  # the words with synonyms not in EDB list

            # check whether the synonyms is in the ones in WordNet
            # the synonyms data[k]
            # the synonyms in mWordNet

            # import pdb; pdb.set_trace()
            #k_wordnet_list = cal.get_wordnet_list(w)
            #k_wordnet_list = wordnet.get_hyponyms(w)
            # k_roget_list = roget.get_roget_synset(w)

            # we should remove the words not in EDB list
            k_simp_wordnet = []
            for wd in k_wordnet_list:
                if wd in wordlist:
                    k_simp_wordnet.append(wd)
                    
            if len(k_simp_wordnet) >= 1:
                _num_not_simp_words += 1  # difficult words with simple substitutions       
            #
            """
            pos_coincolist = data[id][2:]
            coincolist = []
            for pos in pos_coincolist:
                coincolist.append(pos.split('@')[1])
                
            if w in coincolist:
                coincolist.remove(w)
            """
            if w in k_simp_wordnet:
                k_simp_wordnet.remove(w)

            feas = set(coincolist).intersection(k_simp_wordnet)
            if len(feas) >= 1:  #
                num_feasible_words += 1

        # the ceiling of the all words between all words and k_wordnet
        all_feas = set(coincolist).intersection(k_wordnet_list)
        if len(all_feas) >= 1:
            all_num_feasible_words += 1
    #
    print "#num_words: ", num_words
    print "#all_num_feasible_words: ", all_num_feasible_words
    all_ceiling = 0
    if all_num_feasible_words == 0:
        all_ceiling = 0
    else:
        all_ceiling = all_num_feasible_words/float(num_words)

    print "#all_ceiling: ", all_ceiling    
   
    print "#words with feasible: ", num_feasible_words
    print "#num_not_simp_words: ", num_not_simp_words
    print "#_num_not_simp_words: ", _num_not_simp_words
    # import pdb; pdb.set_trace()
    if num_feasible_words == 0:
        ceiling = 0
    else:
        ceiling = num_feasible_words/float(_num_not_simp_words)

    return num_simp_words, _num_not_simp_words, ceiling


def get_semeval_stat_info(filename, synset_filename, store_filename):
    """ store the stat info in store_filename
    """
    # store read the synset file
    synsets = {}  # synsets -> { word : [words]}
    f = open(synset_filename, 'rU')
    for line in f:  # for each line
            line = line.strip('\n')
            # import pdb; pdb.set_trace()
            if line:
                [word, synset] = line.split("::")
                lst = []
                for w in synset.strip().split(";"):
                    if w:
                        w = ''.join(filter(lambda x: not x.isdigit(), w))
                        lst.append(w.rstrip())
                        
                synsets[word.rstrip()] = lst

    f.close()         

    #
    soup = BeautifulSoup(open(filename), "lxml")

    num_sentences = 0
    num_tokens = 0
    num_words_syns = 0

    semeval = OrderedDict()  # semeval -> {side.n 301 : [sent], [synsets]}
    
    lexelts = soup.find_all("lexelt")
    for lex in lexelts:
        # num_words_syns = num_words_syns + 1
        item = lex["item"]
        for sentence in lex.find_all("instance"):
            num_sentences = num_sentences + 1

            id = sentence["id"]
            key = item + " " + id
            sent = sentence.context.get_text()

            # import pdb; pdb.set_trace()
            words = sent.split()
            num_tokens = num_tokens + len(words) - 1

            semeval[key] = []
            semeval[key].append(sent)
            if key in synsets:
                # print(synsets[key])
                semeval[key].append(synsets[key])

    json.dump(semeval, open(store_filename, 'w'))

    """
    with open(store_filename, "w") as outfile:
        json.dump(semeval, outfile)
    
    """
    num_words_syns = num_sentences
    
    return num_sentences, num_tokens, num_words_syns


def get_semeval_info(datafile, wordlist):
    """ 
    """
    num_sentences = 0
    num_tokens = 0
    num_words_syns = 0
    tokens = {}
    #
    num_simp_words = 0  # number of words from datafile and can be found in wordlist
    simp_words = []
    num_not_simp_words = 0
    _num_not_simp_words = 0
    # not_simp_words = []

    num_feasible_words = 0
    feasible_words = []
    
    # the words with synonyms that are not in EDB list
    not_simp_wordlist = []
    _not_simp_wordlist = []

    all_num_feasible_words = 0
    
    # load the dict from semeval dataset
    # data = json.load(open(datafile))
    data = json.load(open(datafile))

    num_words = 0
    for id in data:
        # print data[id] # like "side.n 301"
        # import pdb; pdb.set_trace()
        num_words += 1
        # num_tokens = num_tokens + 1
        # TOUPDATE
        w_pos = id.split()[0]
        w = w_pos.split(".")[0]

        """
        num_sentences += 1
        if w in tokens:
            tokens[w] = tokens[w] + 1
        else:
            tokens[w] = 0
        """

        # import pdb; pdb.set_trace()
        _w = w.lower()
        k_wordnet_list = cal.get_wordnet_list(_w)
            #
        semevallist = []    
        if len(data[id]) > 1:
            semevallist = data[id][1]
            
        if w in semevallist:
            semevallist.remove(w)
            
        if _w in wordlist:  # the lemma
            num_simp_words = num_simp_words + 1
            # simp_words.append(w)
        else:
            if len(k_wordnet_list) >= 1:
                num_not_simp_words = num_not_simp_words + 1
            # not_simp_words.append(w)

            # not_simp_wordlist.append(w) # the words with synonyms not in EDB list

            # check whether the synonyms is in the ones in WordNet
            # the synonyms data[k]
            # the synonyms in mWordNet

            # import pdb; pdb.set_trace()
            # TODO: should check each word in k_wordnet_list is in EDB list
            #k_wordnet_list = cal.get_wordnet_list(_w)
            #k_roget_list = roget.get_roget_synset(_w)

            #lst = k_wordnet_list + k_roget_list
            
            k_simp_wordnet = []
            for wd in k_wordnet_list:
                if wd in wordlist:
                    k_simp_wordnet.append(wd)

            if w in k_simp_wordnet:
                k_simp_wordnet.remove(w)

            if len(k_simp_wordnet) >= 1:
                # _not_simp_wordlist.append(w)
                _num_not_simp_words = _num_not_simp_words + 1

            feas = set(semevallist).intersection(k_simp_wordnet)
            if len(feas) >= 1:  #
                # feasible_words.append(w)
                num_feasible_words = num_feasible_words + 1
                
        # the ceiling of the all words between all words and k_wordnet
        all_feas = set(semevallist).intersection(k_wordnet_list)
        if len(all_feas) >= 1:
            all_num_feasible_words += 1
    #
    print "#num_words: ", num_words
    print "#all_num_feasible_words: ", all_num_feasible_words
    all_ceiling = 0
    if all_num_feasible_words == 0:
        all_ceiling = 0
    else:
        all_ceiling = all_num_feasible_words/float(num_words)

    print "#all_ceiling: ", all_ceiling         
    # num_tokens = len(tokens)
    # num_not_simp_words = len(set(not_simp_wordlist))
    # _num_not_simp_words = len(set(_not_simp_wordlist))
    # num_feasible_words = len(set(feasible_words))
    #
    # print "#num_tokens: ", num_tokens
    print "#num_simp_words: ", num_simp_words
    print "#num_not_simp_words: ", num_not_simp_words
    print "#_num_not_simp_words: ", _num_not_simp_words
    print "#words with feasible: ", num_feasible_words
    # import pdb; pdb.set_trace()
    if num_feasible_words == 0:
        ceiling = 0
    else:
        ceiling = num_feasible_words/float(_num_not_simp_words)

    return num_tokens, num_not_simp_words, _num_not_simp_words, ceiling


def print_semeval_interdata(datafile, wordlist):
    """print the intermedia data for the check
     @datafile is the lemmas dict filename
     @wordlist is the edb wordlist

    """
    data = json.load(open(datafile))

    # import pdb; pdb.set_trace()
    for id in data:
        w_pos = id.split()[0]
        w = w_pos.split(".")[0]
        pos = w_pos.split(".")[1]

        k_wordnet_list = wordcal.get_wordnet_list(w)
        #k_roget_list = roget.get_roget_synset(w)
        # lst = k_wordnet_list + k_roget_list

        # wordlist = k_wordnet_list + k_roget_list

        pos_wordlist = []    
        for w in k_wordnet_list:
            #import pdb; pdb.set_trace()
            _w = wordcal.get_word_pos(str(w))
            pos_wordlist.append(_w)
        
        k_simp_wordnet = []
        for wd in k_wordnet_list:
            if wd in wordlist:
                k_simp_wordnet.append(wd)

        #  import pdb; pdb.set_trace()  #
        if len(data[id]) > 1:
            semevallist = data[id][1]
        if w in semevallist:
            semevallist.remove(w)
            
        if w in k_simp_wordnet:
            k_simp_wordnet.remove(w)

        if w in k_wordnet_list:
            k_wordnet_list.remove(w)

        #if w in k_roget_list:
        #    k_roget_list.remove(w)

        output = OrderedDict()    
        feas = set(semevallist).intersection(k_simp_wordnet)

        wordnet_feas = set(semevallist).intersection(k_wordnet_list)
        #roget_feas = set(semevallist).intersection(k_roget_list)

        if len(feas) >= 1:
            k = '*'+w
            if len(wordnet_feas) >= 1:  # at least one in wordnet
                wf = ['Y']
            else:
                wf = ['N']

            """
            if len(roget_feas) >= 1:
                rf = ['Y']
            else:
                rf = ['N']
            """  
                
            output[k] = [data[id][0],
                         k+"@"+pos,
                         semevallist,
                         pos_wordlist,
                         wf]
        else:
            if len(wordnet_feas) >= 1:  # at least one in wordnet
                wf = ['Y']
            else:
                wf = ['N']

            """
            if len(roget_feas) >= 1:
                rf = ['Y']
            else:
                rf = ['N']
            """
            output[w] = [data[id][0],
                         w+"@"+pos,
                         semevallist,
                         pos_wordlist,
                         wf]

        # write the sentence

        #  import pdb; pdb.set_trace()
        # print(data[id][0])
        #import pdb; pdb.set_trace()
        # wirte the file in CSV format
        with codecs.open('semeval_l1.csv', 'a', encoding='utf-8') as outfile:
            wr = csv.writer(outfile, delimiter = ',', quoting = csv.QUOTE_ALL)
            for row in output:
                wr.writerow(output[row])

        """
        with codecs.open('semeval_l1_.json', 'w', encoding='utf-8') as outfile:
            #outfile.write('\nSentence:'+data[id][0]+'\n')
            json.dump(output, outfile)
        """
    # json.dump(output, open('intermedia.json', 'w'))


def get_mturk_stat_info(filename, store_filename):
    """ store the stat info in store_filename
    """
    num_sentences = 0
    num_words_syns = 0
    num_words = 0
    
    # store read the synset file
    mturk = OrderedDict()  # Synsets -> { word : [sent], [synsets]}

    # this docs is used for print in order
    docs = []

    f = open(filename, 'rU')
    for line in f:  # for each line
        line = line.strip('\n')
        # import pdb; pdb.set_trace()
        if line:
            # import pdb; pdb.set_trace()
            obj = line.split("\t")
            sent = obj[0]
            word = obj[1]
            synsets = obj[2:]

            # docs.append(word) 
            # import pdb; pdb.set_trace()
            num_words = num_words + len(sent.split()) - 1
            num_words_syns = num_words_syns + 1

            w = str(num_words_syns) + '_' + word
            mturk[w] = []
            mturk[w].append(sent)
            mturk[w].append(synsets)

            docs.append(w)
            
            num_sentences = num_sentences + 1
            
    f.close()

    """
    blob = open('unknown-file').read()
    m = magic.Magic(mime_encoding=True)
    encode = m.from_buffer(blob)
    """
    # import pdb; pdb.set_trace()
    json.dump(mturk, open(store_filename, 'w'))
    # json.dump(docs, open('docs.txt', 'w'))

    return num_sentences, num_words, num_words_syns, docs


def get_mturk_info(datafile, wordlist):
    # num_sentences = 0
    num_tokens = 0
    # num_words_syns = 0
    # tokens = {}
    #
    num_simp_words = 0  # number of words from datafile and can be found in wordlist
    # simp_words = []
    num_not_simp_words = 0
    _num_not_simp_words = 0
    # not_simp_words = []

    num_feasible_words = 0
    # feasible_words = []
    
    # the words with synonyms that are not in EDB list
    # not_simp_wordlist = []
    # _not_simp_wordlist = []
    
    # load the dict from semeval dataset
    # data = json.load(open(datafile))
    data = json.load(open(datafile))
    # docs = json.load(open('docs.txt'))

    #num_words = 0
    all_num_feasible_words = 0
    # import pdb; pdb.set_trace()
    # for ind in docs:
    for id in data:
        # print data[id] # 
        # import pdb; pdb.set_trace()
        num_tokens = num_tokens + 1
        # TOUPDATE

        w = wordcal.get_lemma(id.split('_')[1])
        # TODO: should check each word in k_wordnet_list is in EDB list
        k_wordnet_list = wordcal.get_wordnet_list(w)
        #k_roget_list = roget.get_roget_synset(w)

        # import pdb; pdb.set_trace()
        #lst = k_wordnet_list + k_roget_list
            
        # import pdb; pdb.set_trace()
        mturklist = data[id][1]
        if id in mturklist:
            mturklist.remove(id)

        # do the lemma
        _mturklist = []
        for wd in mturklist:
            _mturklist.append(wordcal.get_lemma(wd))
                 
        
        # import pdb; pdb.set_trace()
       
        if w in wordlist:  # the lemma
            num_simp_words += 1
            # simp_words.append(w)
        else:
            num_not_simp_words += 1
            # not_simp_words.append(w)

            # not_simp_wordlist.append(w) # the words with synonyms not in EDB list
            # check whether the synonyms is in the ones in WordNet
            # the synonyms data[k]
            # the synonyms in mWordNet

            # import pdb; pdb.set_trace()
            # TODO: should check each word in k_wordnet_list is in EDB list
            #k_wordnet_list = cal.get_wordnet_list(w)
            #k_roget_list = roget.get_roget_synset(w)

            # import pdb; pdb.set_trace()
            #lst = k_wordnet_list + k_roget_list
            
            k_simp_wordnet=[]
            for wd in k_wordnet_list:
                if wd in wordlist:
                    k_simp_wordnet.append(wd)
              #
            """ 
            # import pdb; pdb.set_trace()
            mturklist = data[id][1]
            if id in mturklist:
                mturklist.remove(id)

            # do the lemma
            _mturklist = []
            for wd in mturklist:
                _mturklist.append(wordcal.get_lemma(wd))
            """    
            if w in k_simp_wordnet:
                k_simp_wordnet.remove(w)
    
            if len(k_simp_wordnet) >= 1:
                # _not_simp_wordlist.append(w)
                _num_not_simp_words = _num_not_simp_words + 1     

            # import pdb; pdb.set_trace()
            feas = set(_mturklist).intersection(k_simp_wordnet)
            if len(feas) >= 1:  #
                # feasible_words.append(w)
                num_feasible_words += 1

        # the ceiling of the all words between all words and k_wordnet
        all_feas = set(_mturklist).intersection(k_wordnet_list)
        if len(all_feas) >= 1:
            all_num_feasible_words += 1
    #
    print "#num_words: ", num_tokens
    print "#all_num_feasible_words: ", all_num_feasible_words
    all_ceiling = 0
    if all_num_feasible_words == 0:
        all_ceiling = 0
    else:
        all_ceiling = all_num_feasible_words/float(num_tokens)

    print "#all_ceiling: ", all_ceiling    

    # num_tokens = len(tokens)
    # num_not_simp_words = len(set(not_simp_wordlist))
    # _num_not_simp_words = len(set(_not_simp_wordlist))
    # num_feasible_words = len(set(feasible_words))
    #
    #print "#num_tokens: ", num_tokens
    print "#num_simp_words: ", num_simp_words
    print "#num_not_simp_words: ", num_not_simp_words
    print "#_num_not_simp_words: ", _num_not_simp_words
    print "#words with feasible: ", num_feasible_words
    # import pdb; pdb.set_trace()
    if num_feasible_words == 0:
        ceiling = 0
    else:
        ceiling = num_feasible_words/float(_num_not_simp_words)

    return num_tokens, num_not_simp_words, _num_not_simp_words, ceiling



def print_mturk_interdata(datafile, wordlist, docs):
    """print the intermedia data for the check
     @datafile is the lemmas dict filename
     @wordlist is the edb wordlist

    """
    data = json.load(open(datafile))

    # import pdb; pdb.set_trace()
    for wd in docs:

        # import pdb; pdb.set_trace()
        id = wd.split('_')[1]
        w = cal.get_lemma(id)
        
        k_wordnet_list = wordcal.get_wordnet_list(w)
        #k_roget_list = roget.get_roget_synset(w)
        #lst = k_wordnet_list+k_roget_list
        pos_wordlist = []    
        for w in k_wordnet_list:
            #import pdb; pdb.set_trace()
            pos_wordlist.append(wordcal.get_word_pos(str(w)))
       
        #wordlist = k_wordnet_list + k_roget_list
        
        k_simp_wordnet=[]
        for _wd in k_wordnet_list:
            if _wd in wordlist:
                k_simp_wordnet.append(_wd)


        # import pdb; pdb.set_trace()#
        mturklist = []
        if len(data[wd][1]) > 1:
            mturklist = data[wd][1]
        if w in mturklist:
            mturklist.remove(w)
            
        if w in k_simp_wordnet:
            k_simp_wordnet.remove(w)

        if w in k_wordnet_list:
            k_wordnet_list.remove(w)

        #if w in k_roget_list:
        #    k_roget_list.remove(w)

        output = OrderedDict()    
        feas = set(mturklist).intersection(k_simp_wordnet)

        wordnet_feas = set(mturklist).intersection(k_wordnet_list)
        #roget_feas = set(mturklist).intersection(k_roget_list)

        if len(feas) >= 1:
            k = '*'+w
            if len(wordnet_feas) >= 1:  # at least one in wordnet
                wf = ['Y']
            else:
                wf = ['N']

            """    
            if len(roget_feas) >= 1:
                rf = ['Y']
            else:
                rf = ['N']
            """   
            output[k] = [data[wd][0],
                         k,
                         mturklist,
                         pos_wordlist,
                         wf]
        else:
            if len(wordnet_feas) >= 1:  # at least one in wordnet
                wf = ['Y']
            else:
                wf = ['N']

            """   
            if len(roget_feas) >= 1:
                rf = ['Y']
            else:
                rf = ['N']
            """
            output[w] = [data[wd][0],
                         w,
                         mturklist,
                         pos_wordlist,
                         wf]

        # write the sentence

        # import pdb; pdb.set_trace()
        # print(data[id][0])
        # import pdb; pdb.set_trace()
        # wirte the file in CSV format
        with codecs.open('mturk_l1.csv', 'a') as outfile:
            wr = csv.writer(outfile, delimiter = ',', quoting = csv.QUOTE_ALL)
            for row in output:
                wr.writerow(output[row])

        """
        with codecs.open('semeval_l1_.json', 'w', encoding='utf-8') as outfile:
            #outfile.write('\nSentence:'+data[id][0]+'\n')
            json.dump(output, outfile)
        """
    # json.dump(output, open('intermedia.json', 'w'))

    
# Main test
def main(): 
    dir = "/Users/zhaowenlong/workspace/proj/dev.nlp/simptext/"

    """
    filename = dir + "dataset/coinco/coinco.xml"
    store_filename = dir + "utils/coinco/coinco_lemmas.txt"

    info = get_stat_info(filename, store_filename)
    print "#sentences: ", info[0]
    print "#words: ", info[1]
    print "#words marked with synonyms: ", info[2]
    # print "words with synonyms: ", info[3]

    #
    xlsx_filename = dir + "dataset/wordlist.xlsx"

    wordlist = read_xlsx_file(xlsx_filename, 1)
 
    info_ = get_coinco_wordlist(store_filename, wordlist)
    print "#words in the EDB list for level 1: ", info_[0]
    print "#words not in the EDB list for level 1:: ", info_[1]
    print "The ceiling for level 1: ", info_[2]

    wordlist = read_xlsx_file(xlsx_filename, 2)
    info_ = get_coinco_wordlist(store_filename, wordlist)
    print "#words in the EDB list for level 1+2: ", info_[0]
    print "#words not in the EDB list for level 1+2:: ", info_[1]
    print "The ceiling for level 1+2:: ", info_[2]
    # the feasible words
    
    wordlist = read_xlsx_file(xlsx_filename, 3)
    info_3 = get_coinco_wordlist(store_filename, wordlist)
    print "#words in the EDB list for level 1+2+3: ", info_3[0]
    print "#words not in the EDB list for level 1+2+3: ", info_3[1]
    print "The ceiling for level 1+2+3:: ", info_3[2]

    wordlist = read_xlsx_file(xlsx_filename, 4)
    info_4 = get_coinco_wordlist(store_filename, wordlist)
    print "#words in the EDB list for level 1+2+3+4: ", info_4[0]
    print "#words not in the EDB list for level 1+2+3+4: ", info_4[1]
    print "The ceiling for level 1+2+3+4: ", info_4[2]
    """
    # print the intermeida data
    #print_intermedia(store_filename, info[3], wordlist)
    
    # print the inter data in the syntactic simplification
    # print_coinco_sent(filename)
    
    """
    # SemEval 2007 - the lexsub_test.xml + key/gold
    filename = dir + "utils/semeval/test/lexsub_test.xml"
    synset_filename = dir + "utils/semeval/key/gold"
    store_filename = dir + "utils/semeval/semeval_synsets.json"
    
    xlsx_filename = dir + "dataset/wordlist.xlsx"
    wordlist = read_xlsx_file(xlsx_filename, 1)

    # store the related info in store_filename
    info = get_semeval_stat_info(filename, synset_filename, store_filename)

    print "#sentences: ", info[0]
    print "#words: ", info[1]
    print "#words marked with synonyms: ", info[2]
    #print "words with synonyms: ", info[3]

    xlsx_filename = dir + "dataset/wordlist.xlsx"

    
    wordlist1 = read_xlsx_file(xlsx_filename, 1)

    info1 = get_semeval_info(store_filename, wordlist1)
    print "#difficult words with substitutions for level 1: ", info1[1]
    print "#difficult words with simple substitutions for level 1:: ", info1[2]
    print "The ceiling for level 1: ", info1[3]
    

    wordlist = read_xlsx_file(xlsx_filename, 2)
    info2 = get_semeval_info(store_filename, wordlist)
 
    #print "words with synonyms: ", info[3]
    print "#difficult words with substitutions for level 1+2: ", info2[1]
    print "#difficult words with simple substitutions for level 1+2: ", info2[2]
    print "The ceiling for level 1+2: ", info2[3]

    
    wordlist = read_xlsx_file(xlsx_filename, 3)
    info3 = get_semeval_info(store_filename, wordlist)
    #print "words with synonyms: ", info[3]
    print "#difficult words with substitutions for level 1+2+3: ", info3[1]
    print "#difficult words with simple substitutions for level 1+2+3: ", info3[2]
    print "The ceiling for level 1+2+3: ", info3[3]

    wordlist = read_xlsx_file(xlsx_filename, 4)
    info_ = get_semeval_info(store_filename, wordlist)
    #print "words with synonyms: ", info[3]
    print "#difficult words with substitutions for level 1+2+3+4: ", info_[1]
    print "#difficult words with simple substitutions for level 1+2+3+4: ", info_[2]
    print "The ceiling for level 1+2+3+4: ", info_[3]
    

    print_semeval_interdata(store_filename, wordlist1)
    
    """
    
    # Mechanical Turk
    filename = dir + "utils/mturk/lex.mturk.txt"
    store_filename = dir + "utils/mturk/mturk_synsets.json"
    xlsx_filename = dir + "dataset/wordlist.xlsx"

    info = get_mturk_stat_info(filename, store_filename)
    print "#sentences: ", info[0]
    print "#words: ", info[1]
    print "#words marked with synonyms: ", info[2]
    #print "#docs[]", info[3]
    
    wordlist1 = read_xlsx_file(xlsx_filename, 1)
    info_ = get_mturk_info(store_filename, wordlist1)
    print "#difficult words with substitutions for level 1: ", info_[1]
    print "#difficult words with simple substitutions for level 1: ", info_[2]
    print "The ceiling for level 1: ", info_[3]
  
    wordlist = read_xlsx_file(xlsx_filename, 2)
    info_ = get_mturk_info(store_filename, wordlist)
    print "#difficult words with substitutions for level 1+2: ", info_[1]
    print "#difficult words with simple substitutions for level 1+2: ", info_[2]
    print "The ceiling for level 1: ", info_[3]

    wordlist = read_xlsx_file(xlsx_filename, 3)
    info_ = get_mturk_info(store_filename, wordlist)
    print "#difficult words with substitutions for level 1+2+3: ", info_[1]
    print "#difficult words with simple substitutions for level 1+2+3: ", info_[2]
    print "The ceiling for level 1: ", info_[3]

    wordlist = read_xlsx_file(xlsx_filename, 4)
    info_ = get_mturk_info(store_filename, wordlist)
    print "#difficult words with substitutions for level 1+2+3: ", info_[1]
    print "#difficult words with simple substitutions for level 1+2+3: ", info_[2]
    print "The ceiling for level 1+2+3: ", info_[3]

    # print inter data - info[3] is docs[]
    print_mturk_interdata(store_filename, wordlist1, info[3]) 
    
    
    """
    lemmas = []
    wd ='mission'
    for pos in roget.parts_of_speech:

        import pdb; pdb.set_trace()
        words = roget.all_entries(wd, pos)
        for w in words:
            lemmas = lemmas + list(w)

    print lemmas
    """

    
if __name__ == '__main__':
    main()
