# -*- coding: utf-8 -*-
#  
"""
   utils.dt
   ~~~~~~~~~~
   data processing

@author wenlong

@TODO: update the software construction

"""

import sys, re, codecs

from bs4 import BeautifulSoup

from collections import OrderedDict
import openpyxl, json, csv

# for Roget's Thesaurus (1911)
import  cal, tool

from app import alg

reload(sys)
sys.setdefaultencoding('utf-8')

def read_xlsx_file(filename, sheetnums):
    """read the xlsx file and stored first sheetnums into words list"""
    # using the openpyxl lib here
    wb = openpyxl.load_workbook(filename)

    #import pdb; pdb.set_trace()
    sheet_names = wb.get_sheet_names()[0:sheetnums]
    #sheet1 = wb.get_sheet_by_name('level 1')
    # sheet2 = wb.get_sheet_by_name('level 2')
    #sheet = wb.get_sheet_by_name(sheets_names)

    # store the simplied words in the words list
    words = []
    for sheet_name in sheet_names:
        worksheet = wb.get_sheet_by_name(sheet_name)
        for x in range(1, worksheet.max_row+1):
            words.append(str(worksheet.cell(row=x, column=1).value).lower())

    # now removing it
    # TODO- the replace function

    #import pdb; pdb.set_trace()
    return tuple(words)


def read_xml_file(filename, word):
    """return the lemmas for the word in filename"""
    lemmas = []
    soup = BeautifulSoup(open(filename))

    # import pdb; pdb.set_trace()
    tokens = soup.find_all("token")
    for tk in tokens:
        # print tk
        if tk['lemma'] == word:
            lemmas.append(tk['lemma'])
            for st in tk.find_all("subst"):
                # print st['lemma']
                lemmas.append(st['lemma'])
    return tuple(lemmas)


def get_stat_info(filename, store_filename):
    """ read the filename and store the words with lemmas in store_filename"""

    num_words_syns = 0

    docs = OrderedDict() # store the info - docs[sentence] = [id,...]
    
    lemmas = {}  # store the word info - lemmas[id] = {word: synsets, ...}
    
    soup = BeautifulSoup(open(filename), "lxml")

    # number of sentences, based on the 'sent' tag
    sentences = soup.find_all("sent")
    num_sentences = len(sentences)

    # number of tokens, based on the 'token' tag
    num_tokens = len(soup.find_all("token"))

    #import pdb; pdb.set_trace()
    for sentence in sentences:
        # the key in docs is the target sentence
        target = str(sentence.targetsentence)
        docs[target] = []
        for tk in sentence.tokens.find_all("token"):
            #print(tk)
            if tk['id'].isdigit():
                num_words_syns += 1

                docs[target].append(tk['id'])

                #test_token[tk['id']]=tk['wordform']
                #_test_token.append(tk['wordform'])
                # the same word are stored in the same slot in the lemmas dict; A better data structure should be used
                
                lemmas[tk['id']] = []
                lemmas[tk['id']].append(tk['wordform'])
                lemmas[tk['id']].append(tk['lemma'])
                for st in tk.find_all("subst"):
                    lemmas[tk['id']].append(st['lemma'])

    #print "#sentence: ", num_sentence
    #print "#words: ", num_words
    #print "#words_syns: ", num_words_syns
    #print "#_test_token: ", len(_test_token)
    #print "#lemmas: ", len(lemmas)

    """
    for lst in _test_token:
        if lst in lemmas:
            pass
        else:
            print "lst: ", lst
    """

    """
    import pdb; pdb.set_trace()
    k = lemmas.keys()      
    k_feas = list(set(k) - set(_test_token))
    print(k_feas)                   
    """
    # write the file

    #write the fiel
    json.dump(lemmas, open(store_filename, 'w'))
    
    return num_sentences, num_tokens, num_words_syns, docs          


def print_intermedia(datafile, docs, wordlist):
    """print the intermedia data for the check
     @datafile is the lemmas dict filename
     @docs is the targetsentence
     @wordlist is the edb wordlist

    """
    data = json.load(open(datafile))

    # import pdb; pdb.set_trace()
    for sentence in docs.keys():
        #print(sentence)
        #print(docs[sentence])
        
        output = OrderedDict()
        for id in docs[sentence]:
            w = data[id][1] # the lemma of the wordform
            # remove the original word
            coincolist = data[id][1:]
            if w in coincolist:
                coincolist.remove(w)
            
            wordlist = cal.get_wordnet_list(w)
            if w in wordlist:
                wordlist.remove(w)

            feas = set(coincolist).intersection(wordlist)
            if len(feas) >= 1:
                k = '*'+data[id][0]
                output[k] = [coincolist, wordlist]
            else:
                output[data[id][0]] = [coincolist, wordlist]

        # write the sentence

        #import pdb; pdb.set_trace()

        with open('coinco_l1.json', 'a') as outfile:
            outfile.write('\n'+sentence+'\n')
            json.dump(output, outfile, indent=2)

    #json.dump(output, open('intermedia.json', 'w'))

def print_syn_sent(filename, sent_file):
    num_sentences = 0
    num_splitted_sentences = 0
    #data = json.load(open(datafile))
    docs = OrderedDict() # store the info - docs[sentence] = [id,...]
    
    soup = BeautifulSoup(open(filename), "lxml")

    # number of sentences, based on the 'sent' tag
    sentences = soup.find_all("targetsentence")
    #num_sentences = len(sentences)

    p = re.compile(r'<.*?>')
    #import pdb; pdb.set_trace()
    output = OrderedDict()
    for sentence in sentences:
        num_sentences = num_sentences + 1
        #print(sentence)
        sent = str(p.sub('', str(sentence)))
        se = re.sub(r'^”|”$', '', sent)
        #sent = str(BeautifulSoup(sentence).text)
        print(se)
        # write the sentence
        #res = ""
        #res = alg.simp_coordi_sent(se)
        #res = alg.simp_subordi_sent(se)
        #res = alg.simp_advcl_sent(se)
        res = alg.simp_parti_sent(se)
        if res: # the
            num_splitted_sentences = num_splitted_sentences + 1
            
        
        output[sentence] = res
        #import pdb; pdb.set_trace()
        
        with open(sent_file, 'a') as outfile:
           outfile.write(str(sentence)+'\n')
           outfile.write("OUTPUT: " + res + '\n')
           outfile.write('-----------------------\n')
           #json.dump(output, outfile, indent=2)

    return num_sentences, num_splitted_sentences       

def print_semeval_sent(filename, sent_file):
    num_sentences = 0
    num_splitted_sentences = 0
    #data = json.load(open(datafile))
    docs = OrderedDict() # store the info - docs[sentence] = [id,...]
    
    soup = BeautifulSoup(open(filename), "lxml")

    # number of sentences, based on the 'sent' tag
    sentences = soup.find_all("instance")
    #num_sentences = len(sentences)

    p = re.compile(r'<.*?>')
    #import pdb; pdb.set_trace()
    output = OrderedDict()
    _output = OrderedDict()
    for sentence in sentences:

        #import pdb; pdb.set_trace()
        num_sentences = num_sentences + 1
        #print(sentence)
        #sent = str(p.sub('', str(sentence)))
        #se = re.sub(r'^”|”$', '', sent)
        se = sentence.context.get_text()
        #sent = str(BeautifulSoup(sentence).text)
        #print(se)
        # write the sentence
        #res = ""
        res = alg.simp_coordi_sent(se)
        #res = alg.simp_subordi_sent(se)
        #res = alg.simp_advcl_sent(se)
        #res = alg.simp_parti_sent(se)
        if res: # the
            num_splitted_sentences = num_splitted_sentences + 1
            
        output[sentence] = res
        #import pdb; pdb.set_trace()

        with open(sent_file, 'a') as outfile:
           outfile.write(str(sentence)+'\n')
           outfile.write("OUTPUT: " + res + '\n')
           outfile.write('-----------------------\n')
           #json.dump(output, outfile, indent=2)

        _output[sentence] = [sentence, res]
        with codecs.open('mturk_sent.csv', 'a') as _outfile:
            wr = csv.writer(_outfile, delimiter = ',', quoting = csv.QUOTE_ALL)
            wr.writerow(_output[sentence])

    return num_sentences, num_splitted_sentences

def print_mturk_sent(filename, sent_file):
    num_sentences = 0
    num_splitted_sentences = 0
    #data = json.load(open(datafile))
    docs = OrderedDict() # store the info - docs[sentence] = [id,...]
    
    #soup = BeautifulSoup(open(filename), "lxml")

    # number of sentences, based on the 'sent' tag
    #sentences = soup.find_all("instance")
    #num_sentences = len(sentences)

    #p = re.compile(r'<.*?>')
    #import pdb; pdb.set_trace()
    output = OrderedDict()
    f = open(filename, 'rU')
    num = 0
    for line in f:
        line = line.strip('\n')
        if line:
            #import pdb; pdb.set_trace()
            num_sentences = num_sentences + 1
            #print(sentence)
            #sent = str(p.sub('', str(sentence)))
            #se = re.sub(r'^”|”$', '', sent)
            #se = sentence.context.get_text()
            #sent = str(BeautifulSoup(sentence).text)
            obj = line.split("\t")
            se = re.sub(r'^"|"$', '', obj[0])
            print(se)
            # write the sentence
            #res = ""
            res = alg.punct.simp_syn_sent_(se)
            #res = alg.simp_syn_sent_(str(se))
            #res = alg.simp_syn_sent_(str(se))
            #res = alg.simp_advcl_sent(str(se))
            #res = alg.simp_parti_sent(str(se))
            #res = alg.simp_adjec_sent(str(se))
            #res = alg.simp_appos_sent(str(se))
            #res = alg.simp_passive_sent(str(re))
            #import pdb; pdb.set_trace()
            print "re: ", re
            if res: # the
                num_splitted_sentences = num_splitted_sentences + 1


            #import pdb; pdb.set_trace()
            output[se] = res
            #import pdb; pdb.set_trace()
        
            with open(sent_file, 'a') as outfile:
                outfile.write(str(se)+'\n')
                outfile.write("OUTPUT: " + res + '\n')
                outfile.write('-----------------------\n')
                #json.dump(output, outfile, indent=2)

            """  
            num = num + 1
            if num == 200:
                break
            """

    return num_sentences, num_splitted_sentences


def get_coinco_wordlist(datafile, wordlist):
    #
    num_simp_words = 0  # number of words from datafile and can be found in wordlist
    num_not_simp_words = 0
    _num_not_simp_words = 0

    num_feasible_words = 0
    
    # the words with synonyms that are not in EDB list
    not_simp_wordlist = []
    
    # load the dict from coinco dataset
    data = json.load(open(datafile))

    #_num = 0
    #import pdb; pdb.set_trace()
    for id in data.keys():
        #print data[k] # the word
        #_num += 1
        # TOUPDATE
        w = data[id][1].lower()
        if w in wordlist: # the lemma
            num_simp_words += 1
        else:
            num_not_simp_words += 1

            not_simp_wordlist.append(data[id][0]) # the words with synonyms not in EDB list

            # check whether the synonyms is in the ones in WordNet
            # the synonyms data[k]
            # the synonyms in mWordNet

            #import pdb; pdb.set_trace()
            k_wordnet_list = utils.wordcal.get_wordnet_list(w)
            #k_roget_list = roget.get_roget_synset(w)

            # we should remove the words not in EDB list

            k_simp_wordnet=[]
            for wd in (k_wordnet_list):
                if wd in wordlist:
                      k_simp_wordnet.append(wd)
                        
            #
            coincolist = data[id][1:]
            if w in coincolist:
                coincolist.remove(w)
            
            if w in k_simp_wordnet:
                k_simp_wordnet.remove(w)

            if len(k_simp_wordnet) >= 1:
                _num_not_simp_words += 1   ## difficult words with simple substitutions           

            feas = set(coincolist).intersection(k_simp_wordnet)
            if len(feas) >= 1: #
                num_feasible_words += 1

    #
    #print "#num: ", _num
    #print "#num_simp_words: ", num_simp_words
    #print "#num_not_simp_words: ", num_not_simp_words
    #print "#words with feasible: ", num_feasible_words
    #import pdb; pdb.set_trace()
    if num_feasible_words == 0:
        ceiling = 0
    else:
        ceiling = num_feasible_words/float(num_not_simp_words)

    return num_simp_words, _num_not_simp_words, ceiling

def cal_mturk_sent(filename):
    f = open(filename, 'rU')
    _num_wrong_output = 0
    num_false = 0
    num_false_positive = 0
    num_false_negative = 0
    num_true = 0
    for line in f:
        line = line.strip('\n')
        if "OUTPUT" in line:
            #import pdb; pdb.set_trace()
            #num_output = num_output + 1
            #print(sentence)
            #sent = str(p.sub('', str(sentence)))
            #se = re.sub(r'^”|”$', '', sent)
            #se = sentence.context.get_text()
            #sent = str(BeautifulSoup(sentence).text)
            ot_flag = line.split(':')[0]
            ot = line.split(':')[1].strip() # having output or not

            #import pdb; pdb.set_trace()
            if ot_flag == "#OUTPUT":  # the wrong classifier
                _num_wrong_output = _num_wrong_output + 1

            if not ot: #don't have the output,consider it as false positive
                num_false = num_false + 1
                if ot_flag == "#OUTPUT":
                    num_false_positive = num_false_positive + 1
            else: # having the output
                num_true = num_true + 1
                if ot_flag == "#OUTPUT": #False Negative
                    num_false_negative = num_false_negative + 1
                    

    #import pdb; pdb.set_trace()
            #match = re.search(r'(^(#OUTPUT):(\w*))', line)
    print "#num_flase: ", num_false
    print "#num_false_positive: ", num_false_positive
    print "#num_false_negative: ", num_false_negative
    print "#num_true: ", num_true
    print "#_num_wrong_output: ", _num_wrong_output
    
    return num_false, num_false_positive, num_false_negative, num_true       


# Main test
def main():
    dir="/Users/zhaowenlong/workspace/proj/dev.nlp/simptext/"

   
    filename = dir + "dataset/coinco/coinco.xml"
    #store_filename = dir + "dataset/coinco/coinco_lemmas.txt"

    """
    info = get_stat_info(filename, store_filename)
    print "#sentences: ", info[0]
    print "#words: ", info[1]
    print "#words marked with synonyms: ", info[2]
    #print "words with synonyms: ", info[3]

    #
    xlsx_filename = dir + "dataset/wordlist.xlsx"

    wordlist = read_xlsx_file(xlsx_filename, 1)
 
    info_ = get_coinco_wordlist(store_filename, wordlist)
    print "#words in the EDB list for level 1: ", info_[0]
    print "#words not in the EDB list for level 1:: ", info_[1]
    print "The ceiling for level 1: ", info_[2]
    
    
    """
    """
    wordlist = read_xlsx_file(xlsx_filename, 2)
    info_ = get_coinco_wordlist(store_filename, wordlist)
    print "#words in the EDB list for level 1+2: ", info_[0]
    print "#words not in the EDB list for level 1+2:: ", info_[1]
    print "The ceiling for level 1+2:: ", info_[2]
    # the feasible words
    
    wordlist = read_xlsx_file(xlsx_filename, 3)
    info_3 = get_coinco_wordlist(store_filename, wordlist)
    print "#words in the EDB list for level 1+2+3: ", info_3[0]
    print "#words not in the EDB list for level 1+2+3: ", info_3[1]
    print "The ceiling for level 1+2+3:: ", info_3[2]

    wordlist = read_xlsx_file(xlsx_filename, 4)
    info_4 = get_coinco_wordlist(store_filename, wordlist)
    print "#words in the EDB list for level 1+2+3+4: ", info_4[0]
    print "#words not in the EDB list for level 1+2+3+4: ", info_4[1]
    print "The ceiling for level 1+2+3+4: ", info_4[2]

    # print the intermeida data
    inter = print_intermedia(store_filename, info[3], wordlist)

    """
    # print the inter data in the syntactic simplification
    #filename = dir + "utils/semeval/test/lexsub_test.xml"
    filename = dir + "utils/mturk/lex.mturk.txt"
    sent_file = dir + "utils/testset/punct_mturk_l1_.md"
    
    _info = print_mturk_sent(filename, sent_file)
    print "Type: Punct Clauses:"
    print "#sentence in mturk: ", _info[0]
    print "#sentence with Syntactic simplification: ", _info[1]
    

    # recall and precision
    #filename = dir + "utils/testset_groundtruth.md"
    #filename = dir + "utils/coordi_mturk_l1_.json"
    #filename = dir + "utils/testset_gt_adverb.md"
    #filename = dir + "utils/testset_gt_appos.md"

    #_info = cal_mturk_sent(filename)
    
    """
    lemmas = []
    wd ='mission'
    for pos in roget.parts_of_speech:

        import pdb; pdb.set_trace()
        words = roget.all_entries(wd, pos)
        for w in words:
            lemmas = lemmas + list(w)

    print lemmas
    """

if __name__ == '__main__':
     main()
