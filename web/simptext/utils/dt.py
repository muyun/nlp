# -*- coding: utf-8 -*-
#  
"""
   utils.dt
   ~~~~~~~~~~
   data processing

@author wenlong
"""

import openpyxl
"""
from collections import OrderedDict

import string
"""
# Read the dataset
# filename = '/Users/zhaowenlong/workspace/proj/dev.nlp/web/simptext/wordlist.xlsx'

"""
Wb = openpyxl.load_workbook(filename)
sheet = wb.get_sheet_by_name('level 1')

#store the simplied words in the simp_words list
simp_words=[]
for x in range(1, sheet.max_row+1):
    simp_words.append(str(sheet.cell(row=x,column=1).value))

# now removing it
#TODO- the replace function
"""

from bs4 import BeautifulSoup

import json

import cal

from collections import OrderedDict

def read_xlsx_file(filename, sheetnums):
    """read the xlsx file and stored first sheetnums into words list"""
    # using the openpyxl lib here
    wb = openpyxl.load_workbook(filename)

    #import pdb; pdb.set_trace()
    sheet_names = wb.get_sheet_names()[0:sheetnums]
    #sheet1 = wb.get_sheet_by_name('level 1')
    # sheet2 = wb.get_sheet_by_name('level 2')
    #sheet = wb.get_sheet_by_name(sheets_names)

    # store the simplied words in the words list
    words = []
    for sheet_name in sheet_names:
        worksheet=wb.get_sheet_by_name(sheet_name)
        for x in range(1, worksheet.max_row+1):
            words.append(str(worksheet.cell(row=x, column=1).value).lower())

    # now removing it
    # TODO- the replace function

    #import pdb; pdb.set_trace()
    return tuple(words)


def read_xml_file(filename, word):
    """return the lemmas for the word in filename"""
    lemmas = []
    soup = BeautifulSoup(open(filename))

    # import pdb; pdb.set_trace()
    tokens = soup.find_all("token")
    for tk in tokens:
        # print tk
        if tk['lemma'] == word:
            lemmas.append(tk['lemma'])
            for st in tk.find_all("subst"):
                # print st['lemma']
                lemmas.append(st['lemma'])
    return tuple(lemmas)


def get_stat_info(filename, store_filename):
    """ read the filename and store the words with lemmas in store_filename"""
    num_sentence = 0
    num_words = 0
    num_words_syns = 0

    lemmas = {}
    
    soup = BeautifulSoup(open(filename), "lxml")

    sentence = soup.find_all("targetsentence")
    num_sentence = len(sentence)

    tokens = soup.find_all("token")
    num_words = len(tokens)

    for tk in tokens:
        if tk['id'].isdigit():
            num_words_syns += 1

            #test_token[tk['id']]=tk['wordform']
            #_test_token.append(tk['wordform'])
            #TODO: the same word are stored in the same slot in the lemmas dict; A better data structure should be used
            #lemmas[tk['wordform']] = []
            lemmas[tk['id']] = []
            lemmas[tk['id']].append(tk['wordform'])
            lemmas[tk['id']].append(tk['lemma'])
            for st in tk.find_all("subst"):
                #
                lemmas[tk['id']].append(st['lemma'])

    #print "#sentence: ", num_sentence
    #print "#words: ", num_words
    #print "#words_syns: ", num_words_syns
    #print "#_test_token: ", len(_test_token)
    #print "#lemmas: ", len(lemmas)

    """
    for lst in _test_token:
        if lst in lemmas:
            pass
        else:
            print "lst: ", lst
    """

    """
    import pdb; pdb.set_trace()
    k = lemmas.keys()      
    k_feas = list(set(k) - set(_test_token))
    print(k_feas)                   
    """
    # write the file

    #import pdb; pdb.set_trace()
    json.dump(lemmas, open(store_filename, 'w'))
    
    return num_sentence, num_words, num_words_syns           


def print_intermedia(datafile, wordlist):
    """print the intermedia data for the check"""
    data = json.load(open(datafile))

    output = {}
    #import pdb; pdb.set_trace()
    for id in data.keys():
        # remove the original word
        w = data[id[1]]
        coincolist = data[id][1:]
        if w in coincolist:
            coincolist.remove(w)
            
        wordlist = cal.get_wordnet_list(w)
        if w in wordlist:
            wordlist.remove(w)

        feas = set(coincolist).intersection(wordlist)
        if len(feas) >= 1: #
            k = '*'+data[id][0]
            output[k]=[coincolist, wordlist]
        else:    
            output[data[id][0]] = [coincolist, wordlist]


    #import pdb; pdb.set_trace()
    with open('intermedia_l1.json', 'w') as outfile:
        json.dump(output, outfile, indent=2)
        
    #json.dump(output, open('intermedia.json', 'w'))

        

def get_simp_wordlist(datafile, wordlist):
    #
    num_simp_words = 0  # number of words from datafile and can be found in wordlist
    num_not_simp_words = 0

    num_feasible_words = 0
    
    # the words with synonyms that are not in EDB list
    not_simp_wordlist = []
    
    # load the dict from coinco dataset
    data = json.load(open(datafile))

    #_num = 0
    #import pdb; pdb.set_trace()
    for id in data.keys():
        #print data[k] # the word
        #_num += 1
        # TOUPDATE
        w = data[id][1]
        if w in wordlist: # the lemma
            num_simp_words += 1
        else:
            num_not_simp_words += 1

            not_simp_wordlist.append(data[id][0]) # the words with synonyms not in EDB list

            # check whether the synonyms is in the ones in WordNet
            # the synonyms data[k]
            # the synonyms in mWordNet
            k_wordnet_list = cal.get_wordnet_list(w)

            #
            coincolist = data[id][1:]
            if w in coincolist:
                coincolist.remove(w)
            
            if w in k_wordnet_list:
                k_wordnet_list.remove(w)

            feas = set(coincolist).intersection(k_wordnet_list)
            if len(feas) >= 1: #
                num_feasible_words += 1

    #
    #print "#num: ", _num
    #print "#num_simp_words: ", num_simp_words
    #print "#num_not_simp_words: ", num_not_simp_words
    #print "#words with feasible: ", num_feasible_words
    #import pdb; pdb.set_trace()
    if num_feasible_words == 0:
        ceiling = 0
    else:
        ceiling = num_feasible_words/float(num_not_simp_words)

    return num_simp_words, num_not_simp_words, ceiling


def cal_ceiling(simp_wordlist):
    """calculate the ceiling"""
    #for word in simp_wordlist:
        # the synonyms in the coinco
        
    return

# Main test
def main():
    dirname="/Users/zhaowenlong/workspace/proj/dev.nlp/web/simptext/"
    filename = dirname + "/dataset/coinco/coinco.xml"
    store_filename = dirname + "/dataset/coinco/lemmas_.txt"
    #info = get_stat_info(filename, store_filename)
    #print "#sentences: ", info[0]
    #print "#words: ", info[1]
    #print "#words marked with synonyms: ", info[2]
    #print "words with synonyms: ", info[3]

    # 
    xlsx_filename = dirname + "/dataset/wordlist.xlsx"
    wordlist = read_xlsx_file(xlsx_filename, 2)
    info_ = get_simp_wordlist(store_filename, wordlist)
    print "#words in the EDB list for level 1+2: ", info_[0]
    print "#words not in the EDB list for level 1+2:: ", info_[1]
    print "The ceiling for level 1+2:: ", info_[2]
    # the feasible words

    """
    wordlist = read_xlsx_file(xlsx_filename, 3)
    info_3 = get_simp_wordlist(store_filename, wordlist)
    print "#words in the EDB list for level 1+2+3: ", info_3[0]
    print "#words not in the EDB list for level 1+2+3: ", info_3[1]
    print "The ceiling for level 1+2+3:: ", info_3[2]

    wordlist = read_xlsx_file(xlsx_filename, 4)
    info_4 = get_simp_wordlist(store_filename, wordlist)
    print "#words in the EDB list for level 1+2+3+4: ", info_4[0]
    print "#words not in the EDB list for level 1+2+3+4: ", info_4[1]
    print "The ceiling for level 1+2+3+4: ", info_4[2]
    """

    # print the intermeida data
    #inter = print_intermedia(store_filename, wordlist)
    

    
if __name__ == '__main__':
    main()
