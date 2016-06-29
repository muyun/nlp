# -*- coding: utf-8 -*-
# coding by wenlong

import openpyxl

#read the dataset
#filename = '/Users/zhaowenlong/workspace/proj/dev.nlp/web/simptext/wordlist.xlsx'

"""
Wb = openpyxl.load_workbook(filename)
sheet = wb.get_sheet_by_name('level 1')

#store the simplied words in the simp_words list
simp_words=[]
for x in range(1, sheet.max_row+1):
    simp_words.append(str(sheet.cell(row=x,column=1).value))

# now removing it
#TODO- the replace function
"""

def read_file(filename):
    """read the xlsx file and stored 1st column into words list"""
    #using the openpyxl lib here
    wb = openpyxl.load_workbook(filename)
    sheet = wb.get_sheet_by_name('level 1')

    #store the simplied words in the words list
    words=[]
    for x in range(1, sheet.max_row+1):
       words.append(str(sheet.cell(row=x,column=1).value))

    # now removing it
    #TODO- the replace function
    return words


def check_word(word, words):
    """ whether the word is in the words list"""
    return
    
# main test
#def main():
    
